"""
Anillo PMM / Trade-Brand pricing engine (for non-Top-100 parts).

Ports the pricing math from the "Anillo PMM & Trade Brand Pricing Tool" workbook
(Pricing Tool sheet). New unit price (cell D67):

    price = MIN( MAX(helper1, min_yoy, min_gm_price), max_yoy )
            * annual_compounding * volume_factor * cust_order_multiplier

Framework auto-selects: Platform Maturity if the part's maturity is known,
else Trade Brand (cell D53).

Reference data is ingested from the workbook's raw sheets (Bookings History,
Cost History, Unique Customers, Control Inputs), cached in memory. The workbook
is confidential and loaded at runtime (path via env PMM_WORKBOOK_PATH).
"""
from __future__ import annotations

import math
import os
import re
import threading
from collections import Counter, defaultdict
from dataclasses import dataclass, field, asdict

import openpyxl

_HERE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # backend/
PMM_WORKBOOK_PATH = os.environ.get("PMM_WORKBOOK_PATH", os.path.join(_HERE, "data", "pmm_workbook.xlsx"))

UNKNOWN_MATURITY = {"", "other / unknown", "unknown", "n/a"}
MODEST_CRES = {"CRES", "HIGH TEMP CRES", "STAINLESS STEEL"}
MODEST_ALU = {"ALUMINUM", "ALUMINUM ALLOY"}

# Base part-number families explicitly designated Modest trade-brand position
# (low-value hardware standards -> 40% GM floor instead of 70%). Matched on the
# NORMALIZED (dashless, upper) part number, so a pattern like NAS1149F covers
# NAS1149F-0863P etc. "XXXX" in the source spec = any trailing chars (a prefix match);
# NAS1149D...H = starts NAS1149D and ends in H.
MODEST_PART_PATTERNS = [
    re.compile(p) for p in (
        r"^NAS1146C",
        r"^NAS1149F",
        r"^MS20002",
        r"^NAS620",
        r"^MS15797",
        r"^MS27183",
        r"^NAS549",
        r"^MS21299",
        r"^NAS1149D.*H$",
    )
]


def norm_part(s) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def is_modest_part(norm_pn: str) -> bool:
    """True if the normalized part number is on the Modest trade-brand list."""
    return any(p.search(norm_pn) for p in MODEST_PART_PATTERNS)


def norm_cust(s) -> str:
    return re.sub(r"\s+", " ", str(s or "").upper()).strip()


@dataclass
class Order:
    date: str
    qty: int
    unit_price: float
    value: float
    customer: str


@dataclass
class PartContext:
    part: str
    found: bool = True
    uom: str = ""
    material: str = ""
    finish: str = ""
    platform: str = ""
    maturity: str = ""
    trade_position: str = ""       # 'Strong Position' | 'Modest Position' | 'N/A - Known Platform'
    framework: str = ""            # 'Trade Brand Strategy' | 'Platform Maturity Strategy'
    anchor_qty: int | None = None
    cost_per_unit: float | None = None
    reference_orders: list[dict] = field(default_factory=list)  # grouped by date+customer
    shipments_2025: dict | None = None          # {units, value, asp, n} — 2025 EA shipments total
    shipment_orders: list[dict] = field(default_factory=list)   # 2025 shipments grouped by date+customer


@dataclass
class PriceBreakdown:
    part: str
    new_unit_price: float | None = None
    new_gross_margin: float | None = None
    governing_rule: str | None = None
    framework: str | None = None
    reference_price: float | None = None
    reference_qty: int | None = None
    reference_customer: str | None = None
    reference_customer_type: str | None = None
    cost_per_unit: float | None = None
    anchor_qty: int | None = None
    min_gm_pct: float | None = None
    helper_min_gm_or_maturity: float | None = None   # D57
    min_yoy_price: float | None = None               # D58
    max_yoy_price: float | None = None               # D59
    min_gm_price: float | None = None                # D60
    annual_compounding: float | None = None          # D62
    volume_factor: float | None = None               # D63
    multiplier: float | None = None                  # D64
    flagged: bool = False
    reason: str = ""


class PMMEngine:
    def __init__(self):
        self._loaded = False
        self._lock = threading.Lock()
        self.parts: dict[str, dict] = {}          # norm_part -> raw aggregation
        self.customers: dict[str, str] = {}       # norm_cust -> type
        self.costs: dict[str, list] = defaultdict(list)  # norm_part -> [(date, norm_cust, unit_cost)]
        self.ci: dict = {}                        # control inputs
        self.ship_2025: dict[str, dict] = {}      # norm_part -> {units, value, n} (2025 EA shipments)
        self.ship_2025_rows: dict[str, list] = defaultdict(list)  # norm_part -> [{date, customer, qty, value}]
        self.customer_rows: list[dict] = []       # [{name, type}] from Unique Customers (display names)
        self.airbus_enabled: set[str] = set()     # norm_cust of Airbus-enabled suppliers

    # ------------------------------------------------------------------ #
    def ensure_loaded(self):
        if self._loaded:
            return
        with self._lock:
            if not self._loaded:
                self._load()
                self._loaded = True

    def _load(self):
        wb = openpyxl.load_workbook(PMM_WORKBOOK_PATH, data_only=True, read_only=True)
        self._load_control_inputs(wb)
        self._load_customers(wb)
        self._load_bookings(wb)
        self._load_costs(wb)
        self._load_shipments_2025(wb)
        self._load_airbus_enabled(wb)

    def _load_airbus_enabled(self, wb):
        if "Airbus Enabled Suppliers" not in wb.sheetnames:
            return
        ws = wb["Airbus Enabled Suppliers"]
        it = ws.iter_rows(values_only=True)
        next(it)  # header
        for r in it:
            if not r:
                continue
            for c in (r[0] if len(r) > 0 else None, r[1] if len(r) > 1 else None):
                if c:
                    self.airbus_enabled.add(norm_cust(c))

    def _load_shipments_2025(self, wb):
        # "Shipped 2015-2025" sheet: total 2025 shipments per part. Units = quanleft (col 12,
        # the true shipped qty — NOT the cumulative 'quantity' col 9), filtered to UoM == EA
        # (col 8). Value = quanleft × unit price (col 7). Year from ship date (col 4), falling
        # back to Year Invoiced (col 23) when the ship date is blank.
        if "Shipped 2015-2025" not in wb.sheetnames:
            return
        ws = wb["Shipped 2015-2025"]
        it = ws.iter_rows(values_only=True)
        next(it)  # header
        agg = self.ship_2025
        for r in it:
            if not r or r[0] is None:
                continue
            shipped = r[4] if len(r) > 4 else None
            yr = getattr(shipped, "year", None)
            if yr is None and len(r) > 23 and isinstance(r[23], (int, float)):
                yr = int(r[23])
            if yr != 2025:
                continue
            if str(r[8] or "").strip().upper() != "EA":
                continue
            ql = r[12] if len(r) > 12 and isinstance(r[12], (int, float)) else None
            if ql is None or ql <= 0:
                continue
            up = r[7] if len(r) > 7 and isinstance(r[7], (int, float)) else 0
            pn = norm_part(r[0])
            g = agg.get(pn)
            if g is None:
                g = agg[pn] = {"units": 0.0, "value": 0.0, "n": 0}
            g["units"] += ql
            g["value"] += ql * (up or 0)
            g["n"] += 1
            self.ship_2025_rows[pn].append({"date": shipped, "customer": str(r[2]).strip() if r[2] else "",
                                            "qty": ql, "value": ql * (up or 0)})

    def _load_control_inputs(self, wb):
        ws = wb["Control Inputs"]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        maturity, trade, altsup, cust_mult, order_mult = {}, {}, {}, {}, {}
        guard = {}
        slopes = {}
        section = None
        for r in rows:
            b = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
            c = r[2] if len(r) > 2 else None
            d = str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
            if not b:
                continue
            if b == "Platform Maturity": section = "maturity"; continue
            if b == "Trade Brand Taxonomy": section = "trade"; continue
            if b == "Market Position - Platform Maturity": section = "alt"; continue
            if b == "Price Increase Guardrails": section = "guard"; continue
            if b == "Volume Dynamics": section = "slope"; continue
            if b == "Customer Type": section = "cust"; continue
            if b == "Order Type": section = "order"; continue
            if b == "Order Minimums": section = "min"; continue
            if not isinstance(c, (int, float)):
                continue
            if section == "maturity": maturity[b] = c
            elif section == "trade": trade[b] = c
            elif section == "alt": altsup[d or b] = c   # key by description (col D) — matches D20 text
            elif section == "guard": guard[b] = c
            elif section == "slope": slopes[b] = c
            elif section == "cust": cust_mult[b] = c
            elif section == "order": order_mult[b] = c
            elif section == "min": guard[b] = c
        self.ci = {
            "maturity_yoy": maturity,     # '1 - Development' -> 0.04 ...
            "trade_min_gm": trade,        # 'Very strong...' 0.7, 'Modest...' 0.4
            "alt_min_gm": altsup,         # description -> 0.65 / 0.55
            "min_yoy": guard.get("Minimum YoY Price Increase %", 0.055),
            "max_yoy": guard.get("Maximum YoY Price Increase %", 0.35),
            "slope_down": slopes.get("Slope for declining volume:", 0.7),
            "slope_up": slopes.get("Slope for increasing volume:", 0.98),
            "cust_mult": cust_mult,       # OEM 1, Distributor 1.33, Tier 1.33
            "order_mult": order_mult,     # OE 1, Spares 2.5
            "dev_yoy": maturity.get("1 - Development", 0.04),
        }

    def _load_customers(self, wb):
        ws = wb["Unique Customers"]
        it = ws.iter_rows(values_only=True)
        next(it)  # header
        for r in it:
            if r and r[0]:
                typ = str(r[2]).strip() if len(r) > 2 and r[2] else "Unknown"
                self.customers[norm_cust(r[0])] = typ
                self.customer_rows.append({"name": str(r[0]).strip(), "type": typ})

    def _load_bookings(self, wb):
        ws = wb["Bookings History"]
        it = ws.iter_rows(values_only=True)
        next(it)  # header
        # cols: A0 Part, C2 Name, G6 Book Date, J9 Quantity, L11 Book Value, M12 U/M,
        #       N13 ASP, P15 Platform, Q16 Maturity, R17 Material, S18 Finish
        agg = self.parts
        for r in it:
            if not r or r[0] is None:
                continue
            pn = norm_part(r[0])
            g = agg.get(pn)
            if g is None:
                g = agg[pn] = {"orig": str(r[0]).strip(), "orders": [], "qtys": [],
                               "uom": "", "material": "", "finish": "", "platform": "", "maturity": ""}
            qty = r[9] if isinstance(r[9], (int, float)) else None
            price = r[13] if isinstance(r[13], (int, float)) else None
            val = r[11] if isinstance(r[11], (int, float)) else None
            date = r[6]
            cust = str(r[2]).strip() if r[2] else ""
            g["orders"].append({"date": date, "qty": qty, "unit_price": price, "value": val, "customer": cust})
            if qty:
                g["qtys"].append(qty)
            # attributes: take first non-empty seen
            for key, idx in (("uom", 12), ("material", 17), ("finish", 18), ("platform", 15), ("maturity", 16)):
                if not g[key] and len(r) > idx and r[idx] not in (None, ""):
                    g[key] = str(r[idx]).strip()

    def _load_costs(self, wb):
        ws = wb["Cost History"]
        it = ws.iter_rows(values_only=True)
        next(it)  # header
        # cols: A0 Inv Date, E4 Name, F5 Part, T19 Unit Cost
        for r in it:
            if not r or len(r) < 20 or r[5] is None:
                continue
            uc = r[19] if isinstance(r[19], (int, float)) else None
            if uc is None:
                continue
            self.costs[norm_part(r[5])].append((r[0], norm_cust(r[4]), uc))

    # ------------------------------------------------------------------ #
    def _derive_trade_position(self, maturity, material, finish, part_norm=""):
        if str(maturity or "").strip().lower() not in UNKNOWN_MATURITY:
            return "N/A - Known Platform"
        # explicit modest-part list wins over the material/finish heuristic
        if part_norm and is_modest_part(part_norm):
            return "Modest Position"
        mat = str(material or "").strip().upper()
        fin = str(finish or "").strip().upper()
        if (mat in MODEST_CRES and fin == "PASSIVATED") or (mat in MODEST_ALU and fin == "NONE"):
            return "Modest Position"
        return "Strong Position"

    def _cost_for(self, pn, customer_norm):
        rows = self.costs.get(pn)
        if not rows:
            return None
        pc = [r for r in rows if r[1] == customer_norm]
        pick = pc if pc else rows
        # most recent by date
        pick = sorted(pick, key=lambda x: (x[0] is not None, x[0]))
        return pick[-1][2] if pick else None

    def booking_asp_max(self, part: str):
        """Highest per-line ASP ever booked for this part (anomaly-guard reference).
        Returns None if the part has no booking history."""
        self.ensure_loaded()
        g = self.parts.get(norm_part(part))
        if not g:
            return None
        asps = [o["unit_price"] for o in g["orders"]
                if isinstance(o["unit_price"], (int, float)) and o["unit_price"] > 0]
        return max(asps) if asps else None

    def classify(self, part: str):
        """Pricing category for a non-contract part:
        'platform_maturity' (known platform) | 'modest_trade_brand' | 'strong_trade_brand'.
        Returns None if the part has no booking history to classify from."""
        self.ensure_loaded()
        pn = norm_part(part)
        g = self.parts.get(pn)
        if not g:
            return None
        maturity = g["maturity"]
        if str(maturity or "").strip().lower() not in UNKNOWN_MATURITY:
            return "platform_maturity"
        tp = self._derive_trade_position(maturity, g["material"], g["finish"], pn)
        return "modest_trade_brand" if tp == "Modest Position" else "strong_trade_brand"

    def _shipments_2025(self, pn: str):
        s = self.ship_2025.get(pn)
        if not s or s["units"] <= 0:
            return None
        return {"units": round(s["units"]), "value": round(s["value"], 2),
                "asp": round(s["value"] / s["units"], 6) if s["value"] else None, "n": s["n"]}

    def _shipment_orders(self, pn: str):
        """2025 shipments grouped by (date, customer) -> value-weighted unit price,
        same shape as booking reference_orders so the UI can render one shared table."""
        rows = self.ship_2025_rows.get(pn)
        if not rows:
            return []
        grp = {}
        for o in rows:
            key = (str(o["date"]), o["customer"])
            e = grp.setdefault(key, {"qty": 0.0, "value": 0.0, "customer": o["customer"], "date": o["date"]})
            e["qty"] += o["qty"] or 0
            e["value"] += o["value"] or 0
        out = []
        for e in grp.values():
            if e["qty"] > 0:
                out.append({"date": (e["date"].isoformat() if hasattr(e["date"], "isoformat") else str(e["date"])),
                            "customer": e["customer"], "qty": int(e["qty"]),
                            "unit_price": round(e["value"] / e["qty"], 6) if e["value"] else None,
                            "customer_type": self.customers.get(norm_cust(e["customer"]), "Unknown")})
        out.sort(key=lambda x: x["date"], reverse=True)
        return out

    def part_context(self, part: str) -> PartContext:
        self.ensure_loaded()
        pn = norm_part(part)
        ship = self._shipments_2025(pn)
        ship_orders = self._shipment_orders(pn)
        g = self.parts.get(pn)
        if not g:
            return PartContext(part=part, found=False, shipments_2025=ship, shipment_orders=ship_orders)
        maturity = g["maturity"]
        tp = self._derive_trade_position(maturity, g["material"], g["finish"], pn)
        framework = ("Trade Brand Strategy"
                     if str(maturity or "").strip().lower() in UNKNOWN_MATURITY
                     else "Platform Maturity Strategy")
        anchor = None
        if g["qtys"]:
            anchor = Counter(g["qtys"]).most_common(1)[0][0]  # MODE (fallback avg)
            if list(g["qtys"]).count(anchor) == 1:
                anchor = round(sum(g["qtys"]) / len(g["qtys"]))
        # group orders by (date, customer) -> value-weighted price
        grp = defaultdict(lambda: {"qty": 0.0, "value": 0.0, "customer": "", "date": None})
        for o in g["orders"]:
            key = (str(o["date"]), o["customer"])
            e = grp[key]
            e["customer"] = o["customer"]; e["date"] = o["date"]
            if o["qty"]: e["qty"] += o["qty"]
            if o["value"]: e["value"] += o["value"]
            elif o["qty"] and o["unit_price"]: e["value"] += o["qty"] * o["unit_price"]
        refs = []
        for e in grp.values():
            if e["qty"] > 0 and e["value"] > 0:
                refs.append({"date": (e["date"].isoformat() if hasattr(e["date"], "isoformat") else str(e["date"])),
                             "customer": e["customer"], "qty": int(e["qty"]),
                             "unit_price": round(e["value"] / e["qty"], 6),
                             "customer_type": self.customers.get(norm_cust(e["customer"]), "Unknown")})
        refs.sort(key=lambda x: x["date"], reverse=True)
        cost = self._cost_for(pn, norm_cust(refs[0]["customer"]) if refs else "")
        return PartContext(
            part=g["orig"], found=True, uom=g["uom"], material=g["material"], finish=g["finish"],
            platform=g["platform"], maturity=maturity or "Unknown", trade_position=tp, framework=framework,
            anchor_qty=anchor, cost_per_unit=cost, reference_orders=refs, shipments_2025=ship,
            shipment_orders=ship_orders,
        )

    def _min_gm(self, ctx: PartContext, supplier_dynamic: str | None) -> float | None:
        if ctx.trade_position == "Modest Position":
            return self.ci["trade_min_gm"].get("Modest trade brand position", 0.4)
        if ctx.trade_position == "Strong Position":
            return self.ci["trade_min_gm"].get("Very strong trade brand position", 0.7)
        # known platform -> alt supplier dynamic
        if supplier_dynamic and supplier_dynamic in self.ci["alt_min_gm"]:
            return self.ci["alt_min_gm"][supplier_dynamic]
        # default to the more conservative if unspecified
        return min(self.ci["alt_min_gm"].values()) if self.ci["alt_min_gm"] else 0.55

    def price(self, part: str, *, reference_price: float, reference_qty: int | None,
              reference_customer_type: str, new_customer_type: str, order_type: str,
              new_qty: int, trade_position: str, maturity: str, anchor_qty: int | None,
              cost_per_unit: float | None, supplier_dynamic: str | None = None,
              years_since_ref: int = 0) -> PriceBreakdown:
        self.ensure_loaded()
        ci = self.ci
        bd = PriceBreakdown(part=part, reference_price=reference_price, reference_qty=reference_qty,
                            reference_customer_type=reference_customer_type, cost_per_unit=cost_per_unit,
                            anchor_qty=anchor_qty)
        if reference_price is None:
            bd.flagged = True; bd.reason = "No reference order price"; return bd

        known = str(maturity or "").strip().lower() not in UNKNOWN_MATURITY
        framework = "Platform Maturity Strategy" if known else "Trade Brand Strategy"
        bd.framework = framework

        # min GM %
        ctx_stub = PartContext(part=part, trade_position=trade_position)
        min_gm = self._min_gm(ctx_stub, supplier_dynamic)
        bd.min_gm_pct = min_gm

        # helper1 (D57)
        if framework == "Platform Maturity Strategy":
            mat_yoy = ci["maturity_yoy"].get(maturity, ci["dev_yoy"])
            helper1 = (mat_yoy + 1) * reference_price
        else:
            helper1 = (cost_per_unit / (1 - min_gm)) if (cost_per_unit is not None and min_gm < 1) else 0.0
        bd.helper_min_gm_or_maturity = round(helper1, 6)

        # D58 min YoY, D59 max YoY, D60 min-GM price
        min_yoy = reference_price * (1 + (ci["dev_yoy"] if framework == "Platform Maturity Strategy" else ci["min_yoy"]))
        max_yoy = reference_price * (1 + ci["max_yoy"])
        min_gm_price = (cost_per_unit / (1 - min_gm)) if (framework == "Platform Maturity Strategy" and cost_per_unit is not None and min_gm < 1) else 0.0
        bd.min_yoy_price = round(min_yoy, 6); bd.max_yoy_price = round(max_yoy, 6); bd.min_gm_price = round(min_gm_price, 6)

        # D62 compounding
        comp = (1 + ci["min_yoy"]) ** max(years_since_ref - 1, 0) if years_since_ref > 1 else 1.0
        bd.annual_compounding = round(comp, 6)

        # D63 volume factor
        vf = 1.0
        if anchor_qty and new_qty and anchor_qty > 0 and new_qty > 0:
            lg = math.log(new_qty / anchor_qty, 2)
            vf = max(ci["slope_down"] ** lg, ci["slope_up"] ** lg)
        bd.volume_factor = round(vf, 6)

        # D64 customer/order multiplier
        cm = ci["cust_mult"]; om = ci["order_mult"]
        new_cm = cm.get(new_customer_type, 1.0); ref_cm = cm.get(reference_customer_type, 1.0)
        mult = (new_cm / ref_cm if ref_cm else 1.0) * om.get(order_type, 1.0)
        bd.multiplier = round(mult, 6)

        # D67 governing
        helpers = [("Minimum GM% Trade Brand / Platform Maturity Increase", helper1),
                   ("Minimum Price Increase per Year", min_yoy),
                   ("Minimum Gross Margin %", min_gm_price)]
        governing_val = max(h[1] for h in helpers)
        governing_name = next(h[0] for h in helpers if h[1] == governing_val)
        capped = min(governing_val, max_yoy)
        if capped == max_yoy and max_yoy < governing_val:
            governing_name = "Maximum Price Increase per Year"
        price = capped * comp * vf * mult
        bd.new_unit_price = round(price, 6)
        bd.governing_rule = governing_name
        if cost_per_unit is not None and price:
            bd.new_gross_margin = round((price - cost_per_unit) / price, 6)
        return bd


engine = PMMEngine()
