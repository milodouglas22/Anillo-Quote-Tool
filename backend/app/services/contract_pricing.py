"""
Contract pricing for the Anillo Quote Tool.

Aligned to the Price List ("pricing tool") logic so a quote prices contract parts
identically:

  Universe = every part on a Boeing OR Airbus contract:
    - Boeing price = accepted New Boeing Offer, Bucket 6 (floor).
    - Airbus price = highest Anillo contract 'Current Price' among Airbus-family customers.
  Anchor (baseline) = higher of the Boeing / Airbus price.

  Pricing a (part, customer, qty):
    1. part not in the contract universe        -> out_of_scope (goes to the pmm path)
    2. customer is on the part's contract        -> that contract price
                                                    (Boeing-family -> the accepted offer price)
    3. otherwise (non-contract customer)         -> anchor x 1.40 (qty >= 50k) / x 1.60 (qty < 50k)

Customer matching is exact-then-prefix tolerant so truncated/variant feed names
(e.g. "BOEING COMMERCIAL AIRPLA" -> "BOEING COMMERCIAL AIRPLANES") still resolve,
mirroring the Price List fix. Data files load lazily and cache in memory.
"""
from __future__ import annotations

import os
import re
import threading
from dataclasses import dataclass

import openpyxl

_HERE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # backend/
CONTRACT_DB_PATH = os.environ.get("CONTRACT_DB_PATH", os.path.join(_HERE, "data", "contract_db.xlsx"))
BOEING_OFFER_PATH = os.environ.get("BOEING_OFFER_PATH", os.path.join(_HERE, "data", "boeing_offer.xlsx"))
CONTRACT_SITE = os.environ.get("CONTRACT_SITE", "Anillo")

QTY_THRESHOLD = 50_000
MARKUP_HIGH_VOL = 1.40   # qty >= threshold
MARKUP_LOW_VOL = 1.60    # qty <  threshold
BOEING_OFFER_BUCKET_COL = 17   # "Bucket Price 6" = the accepted floor (anchor for non-contract markup)
# Boeing offer volume tiers: "Max Bucket Quantity 1..10" in cols 2..11, "Bucket Price 1..10" in cols 12..21.
BOEING_MAXQTY_COL0 = 2         # first Max Bucket Quantity column
BOEING_PRICE_COL0 = 12         # first Bucket Price column
BOEING_N_BUCKETS = 10


def norm_part(s) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def norm_cust(s) -> str:
    # Whitespace + case only (NO suffix stripping) — matches the validated Price List behavior
    # and avoids over-collapsing distinct customers (e.g. "ACME INC" vs "ACME CO").
    return re.sub(r"\s+", " ", str(s or "").upper()).strip()


def family_of(customer_name) -> str | None:
    u = str(customer_name or "").upper()
    if "BOEING" in u:
        return "Boeing"
    if "AIRBUS" in u:
        return "Airbus"
    if "SPIRIT" in u:
        return "Spirit"
    return None


def contract_price(contract: dict, nc: str):
    """Customer's contract price, tolerating truncated/variant names (prefix match, length-guarded)."""
    p = contract.get(nc)
    if p is not None:
        return p
    if len(nc) >= 8:
        for k, v in contract.items():
            if k.startswith(nc) or nc.startswith(k):
                return v
    return None


def _load_boeing_offer():
    """Returns (floor, buckets):
      floor[pn]   = Bucket Price 6 (col 17) — the accepted floor / non-contract anchor (unchanged).
      buckets[pn] = [(max_qty, price), ...] for buckets 1..10, ascending by max_qty — the
                    volume-tiered contract price a Boeing customer actually pays."""
    if not os.path.exists(BOEING_OFFER_PATH):
        return {}, {}
    wb = openpyxl.load_workbook(BOEING_OFFER_PATH, data_only=True, read_only=True)
    ws = wb["New Boeing Offer"] if "New Boeing Offer" in wb.sheetnames else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    next(it); next(it)   # two header rows
    floor, buckets = {}, {}
    for r in it:
        if not r or r[0] is None:
            continue
        pn = norm_part(r[0])
        p6 = r[BOEING_OFFER_BUCKET_COL] if len(r) > BOEING_OFFER_BUCKET_COL else None
        if isinstance(p6, (int, float)):
            floor[pn] = round(float(p6), 6)
        bk = []
        for i in range(BOEING_N_BUCKETS):
            mq = r[BOEING_MAXQTY_COL0 + i] if len(r) > BOEING_MAXQTY_COL0 + i else None
            pr = r[BOEING_PRICE_COL0 + i] if len(r) > BOEING_PRICE_COL0 + i else None
            if isinstance(mq, (int, float)) and isinstance(pr, (int, float)):
                bk.append((float(mq), round(float(pr), 6)))
        if bk:
            bk.sort(key=lambda x: x[0])
            buckets[pn] = bk
    wb.close()
    return floor, buckets


def boeing_bucket_price(buckets, qty):
    """Volume-matched Boeing bucket price: first bucket whose Max Bucket Quantity >= qty
    (buckets ascending). Above all tiers -> the highest-volume (floor) price."""
    if not buckets:
        return None
    q = qty if isinstance(qty, (int, float)) and qty > 0 else None
    if q is None:
        return buckets[-1][1]   # no qty -> floor
    for max_qty, price in buckets:
        if q <= max_qty:
            return price
    return buckets[-1][1]


@dataclass
class PriceResult:
    in_scope: bool
    flagged: bool
    rule: str                # out_of_scope | contract | markup_40 | markup_60 | no_contract
    unit_price: float | None
    baseline: float | None = None
    matched_customer: str | None = None
    reason: str = ""


class PricingEngine:
    def __init__(self):
        self._loaded = False
        self._lock = threading.Lock()
        self.boeing: dict[str, float] = {}                     # norm_part -> accepted offer floor (Bucket 6)
        self.boeing_buckets: dict[str, list] = {}              # norm_part -> [(max_qty, price)] volume tiers
        self.contract: dict[str, dict[str, float]] = {}        # norm_part -> {norm_cust: price}
        self.family_price: dict[str, dict[str, float]] = {}    # norm_part -> {family: max price}
        self.baseline_map: dict[str, float] = {}               # norm_part -> anchor (higher of Boeing/Airbus)
        self.contract_parts: set[str] = set()                  # the universe (Boeing offer ∪ Airbus contract)
        self.raw_by_norm: dict[str, set] = {}                  # norm_part -> {raw part strings seen} (collision guard)
        self.customers: list[str] = []                         # distinct contract customer display names

    def ensure_loaded(self):
        if self._loaded:
            return
        with self._lock:
            if self._loaded:
                return
            self._load()
            self._loaded = True

    def _load(self):
        self.boeing, self.boeing_buckets = _load_boeing_offer()

        wb = openpyxl.load_workbook(CONTRACT_DB_PATH, data_only=True, read_only=True)
        ws = wb["Consolidated Contracts"] if "Consolidated Contracts" in wb.sheetnames else wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        hdr = [(" ".join(str(h).split()) if h else "") for h in next(it)]

        def find(prefix):
            return next((i for i, h in enumerate(hdr) if h.startswith(prefix)), None)

        i_part, i_price, i_cust = find("Part Number"), find("Current Price"), find("Customer Name (Site)")
        cust_set = set()
        for r in it:
            if r[i_part] is None or str(r[1]).strip() != CONTRACT_SITE:
                continue
            price = r[i_price] if isinstance(r[i_price], (int, float)) else None
            if price is None:
                continue
            pn = norm_part(r[i_part])
            self.raw_by_norm.setdefault(pn, set()).add(str(r[i_part]).strip())
            cust = str(r[i_cust]).strip() if r[i_cust] else ""
            self.contract.setdefault(pn, {})[norm_cust(cust)] = price
            fam = family_of(cust)
            if fam and price > self.family_price.setdefault(pn, {}).get(fam, 0):
                self.family_price[pn][fam] = price
            if cust:
                cust_set.add(cust)
        wb.close()

        # Boeing-family prices come from the accepted offer (overrides the contract-DB current price)
        for pn, bp in self.boeing.items():
            if pn in self.contract:
                for nc in list(self.contract[pn]):
                    if family_of(nc) == "Boeing":
                        self.contract[pn][nc] = bp
            self.family_price.setdefault(pn, {})["Boeing"] = bp

        # universe + anchor (higher of Boeing offer / Airbus contract)
        self.contract_parts = set(self.boeing) | {pn for pn, fp in self.family_price.items() if fp.get("Airbus") is not None}
        for pn in self.contract_parts:
            fp = self.family_price.get(pn, {})
            cand = [p for p in (self.boeing.get(pn), fp.get("Airbus")) if p is not None]
            if cand:
                self.baseline_map[pn] = max(cand)
        self.customers = sorted(cust_set)

    def is_contract(self, part) -> bool:
        self.ensure_loaded()
        return norm_part(part) in self.contract_parts

    def contract_info(self, part):
        """For a contract-universe part: the contract (anchor) price and which contract it
        references (Boeing / Airbus / Both). None if the part isn't on a contract."""
        self.ensure_loaded()
        pn = norm_part(part)
        if pn not in self.contract_parts:
            return None
        fp = self.family_price.get(pn, {})
        boeing = self.boeing.get(pn, fp.get("Boeing"))
        airbus = fp.get("Airbus")
        baseline = self.baseline_map.get(pn)
        cands = [(f, p) for f, p in (("Boeing", boeing), ("Airbus", airbus)) if p is not None]
        if len(cands) == 2 and abs(cands[0][1] - cands[1][1]) < 1e-9:
            family = "Boeing & Airbus"
        elif cands:
            family = max(cands, key=lambda x: x[1])[0]
        else:
            family = None
        return {"contract_price": baseline, "family": family,
                "boeing_price": boeing, "airbus_price": airbus}

    def price(self, part: str, customer: str, qty) -> PriceResult:
        self.ensure_loaded()
        pn = norm_part(part)
        if pn not in self.contract_parts:
            return PriceResult(False, False, "out_of_scope", None, reason="Not on a Boeing/Airbus contract")
        baseline = self.baseline_map.get(pn)
        nc = norm_cust(customer)
        fam = family_of(customer)
        q = qty if isinstance(qty, (int, float)) else 0
        # Boeing customer on a Boeing-offer part -> the bucket price for the ordered VOLUME
        # (NOT the Bucket-6 floor, which is only the anchor for non-contract markup).
        if fam == "Boeing" and pn in self.boeing_buckets:
            bp = boeing_bucket_price(self.boeing_buckets[pn], q)
            if bp is not None:
                return PriceResult(True, False, "contract", round(bp, 6), baseline=baseline,
                                   matched_customer=customer, reason="Boeing bucket price (volume-matched)")
        # other on-contract customer -> the contract price
        cp = contract_price(self.contract.get(pn, {}), nc)
        if cp is None and fam == "Boeing" and pn in self.boeing:
            cp = self.boeing[pn]
        if cp is None and fam == "Airbus":
            cp = self.family_price.get(pn, {}).get("Airbus")
        if cp is not None:
            return PriceResult(True, False, "contract", round(cp, 6), baseline=baseline,
                               matched_customer=customer, reason="On-contract customer price")
        if baseline is None:
            return PriceResult(True, True, "no_contract", None, reason="No baseline contract price")
        # non-contract customer -> anchor x volume markup
        q = qty if isinstance(qty, (int, float)) else 0
        markup = MARKUP_HIGH_VOL if q >= QTY_THRESHOLD else MARKUP_LOW_VOL
        rule = "markup_40" if q >= QTY_THRESHOLD else "markup_60"
        return PriceResult(True, False, rule, round(baseline * markup, 6), baseline=baseline,
                           reason=f"Anchor {baseline} x {markup}")


engine = PricingEngine()
