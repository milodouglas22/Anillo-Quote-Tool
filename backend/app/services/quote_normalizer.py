"""
Quote normalizer: turn any incoming customer RFQ/quote file (Excel or PDF) into
Anillo's standard reply format.

Reply format (one row per part):
    Part Number | Qty 1 | Price 1 | Qty 2 | Price 2 | Qty 3 | Price 3 | L/T | MFG | REV

Known formats are handled by dedicated adapters. Unknown formats fall back to a
generic parser that returns the raw columns + a predicted column mapping for the
user to confirm (manual column-matching exercise).
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from rapidfuzz import fuzz

REPLY_COLUMNS = [
    "Part Number", "Qty 1", "Price 1", "Qty 2", "Price 2",
    "Qty 3", "Price 3", "L/T", "MFG", "REV",
]

# Which reply columns a user must map for a generic file to be usable.
REQUIRED_REPLY_COLUMNS = ["Part Number", "Qty 1"]


@dataclass
class NormalizeResult:
    filename: str
    detected_format: str            # 'adept' | 'incora' | 'boeing_sap' | 'boeing_pdf' | 'unknown'
    recognized: bool
    rows: list[dict] = field(default_factory=list)      # reply-format rows (when recognized)
    warnings: list[str] = field(default_factory=list)
    customer_guess: str = ""                              # best-effort customer parsed from the quote
    # For unknown formats -> data the manual column-matcher needs:
    source_columns: list[str] = field(default_factory=list)
    sample_data: dict = field(default_factory=dict)
    candidate_rows: list[list] = field(default_factory=list)
    header_row_index: int = 0
    suggestions: dict = field(default_factory=dict)     # reply_col -> [{source_col, score}]
    raw_records: list[dict] = field(default_factory=list)  # every source row as {col: val}

    def to_dict(self) -> dict:
        return {
            "filename": self.filename,
            "detected_format": self.detected_format,
            "recognized": self.recognized,
            "rows": self.rows,
            "warnings": self.warnings,
            "customer_guess": self.customer_guess,
            "source_columns": self.source_columns,
            "sample_data": self.sample_data,
            "candidate_rows": self.candidate_rows,
            "header_row_index": self.header_row_index,
            "suggestions": self.suggestions,
            "raw_records": self.raw_records,
            "reply_columns": REPLY_COLUMNS,
        }


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def _s(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v):
    """Coerce to number if it looks numeric, else return cleaned string/None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    txt = str(v).strip().replace(",", "")
    if txt == "":
        return None
    try:
        f = float(txt)
        return int(f) if f.is_integer() else f
    except ValueError:
        return str(v).strip()


def _load_sheet_matrix(content: bytes) -> tuple[str, list[list]]:
    """Return (sheet_name, rows-as-lists-of-values) for the first non-empty sheet."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    for ws in wb.worksheets:
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        # strip fully-empty trailing rows
        rows = [r for r in rows if any(_s(c) for c in r)]
        if rows:
            return ws.title, rows
    return (wb.worksheets[0].title if wb.worksheets else "Sheet"), []


def _find_row(rows: list[list], *tokens: str) -> int:
    """Index of the first row containing ALL given tokens (case-insensitive)."""
    toks = [t.lower() for t in tokens]
    for i, row in enumerate(rows):
        cells = [_s(c).lower() for c in row]
        joined = " | ".join(cells)
        if all(any(t == c or t in c for c in cells) or t in joined for t in toks):
            return i
    return -1


def _empty_reply_row() -> dict:
    return {c: None for c in REPLY_COLUMNS}


# --------------------------------------------------------------------------- #
# Format detection
# --------------------------------------------------------------------------- #
def detect_format(filename: str, content: bytes) -> str:
    name = filename.lower()
    if name.endswith(".pdf"):
        head = content[:4096].decode("latin-1", errors="ignore")
        # (PDF text is compressed; we detect Boeing by fuller extraction later too)
        return "boeing_pdf"
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        try:
            _, rows = _load_sheet_matrix(content)
        except Exception:
            return "unknown"
        flat = " | ".join(_s(c).lower() for r in rows[:15] for c in r)
        if "rfq_number" in flat and "material_number" in flat and "scale_qty1" in flat:
            return "boeing_sap"
        if "adept part" in flat and "qty 1" in flat:
            return "adept"
        if "incora quote ref" in flat or ("customer part" in flat and "quote qty 1" in flat):
            return "incora"
    return "unknown"


# --------------------------------------------------------------------------- #
# Adapters for known formats
# --------------------------------------------------------------------------- #
def _adapt_adept(rows: list[list], res: NormalizeResult) -> None:
    res.customer_guess = "Adept"   # ADEPT-format quotes come from the customer Adept
    hi = _find_row(rows, "adept part")
    header = [_s(c).lower() for c in rows[hi]]

    def col(*names):
        for n in names:
            for j, h in enumerate(header):
                if h == n:
                    return j
        return None

    c_part = col("adept part")
    c = {
        "Qty 1": col("qty 1"), "Price 1": col("price 1"),
        "Qty 2": col("qty 2"), "Price 2": col("price 2"),
        "Qty 3": col("qty 3"), "Price 3": col("price 3"),
        "L/T": col("del", "l/t", "lead time"), "MFG": col("mfg"), "REV": col("rev"),
    }
    for row in rows[hi + 1:]:
        part = _s(row[c_part]) if c_part is not None and c_part < len(row) else ""
        if not part:
            continue
        rr = _empty_reply_row()
        rr["Part Number"] = part
        for key, j in c.items():
            if j is not None and j < len(row):
                rr[key] = _num(row[j])
        res.rows.append(rr)


def _adapt_incora(rows: list[list], res: NormalizeResult) -> None:
    # The quote is FROM Incora, so the customer we're quoting to is Incora
    # (the "Customer" field inside the sheet is Incora's own end customer — not ours).
    res.customer_guess = "Incora"
    hi = _find_row(rows, "customer part")
    header = [_s(c).lower() for c in rows[hi]]

    def col(name):
        for j, h in enumerate(header):
            if h == name:
                return j
        return None

    c_part = col("customer part")
    c_quoted = col("quoted part number")
    tiers = [(col("quote qty 1"), col("cost 1")),
             (col("quote qty 2"), col("cost 2")),
             (col("quote qty 3"), col("cost 3"))]
    c_lt = col("lead time")
    has_t4 = col("quote qty 4") is not None
    if has_t4:
        res.warnings.append("Incora file has a 4th quantity break; reply format holds only 3 (4th dropped).")

    for row in rows[hi + 1:]:
        part = _s(row[c_part]) if c_part is not None and c_part < len(row) else ""
        quoted = _s(row[c_quoted]) if c_quoted is not None and c_quoted < len(row) else ""
        part = part or quoted
        if not part:
            continue
        rr = _empty_reply_row()
        rr["Part Number"] = part
        for i, (cq, cc) in enumerate(tiers, start=1):
            if cq is not None and cq < len(row):
                rr[f"Qty {i}"] = _num(row[cq])
            if cc is not None and cc < len(row):
                rr[f"Price {i}"] = _num(row[cc])
        if c_lt is not None and c_lt < len(row):
            rr["L/T"] = _num(row[c_lt])
        res.rows.append(rr)


def _adapt_boeing_sap(rows: list[list], res: NormalizeResult) -> None:
    hi = _find_row(rows, "rfq_number", "material_number")
    header = [_s(c).lower() for c in rows[hi]]

    def col(prefix):
        for j, h in enumerate(header):
            if h.startswith(prefix):
                return j
        return None

    c_part = col("material_number")
    c_rfqqty = col("rfq_quantity")
    tiers = [(col("scale_qty1"), col("scale_price1"), col("scale_leadtime1")),
             (col("scale_qty2"), col("scale_price2"), col("scale_lead_time2")),
             (col("scale_qty3"), col("scale_price3"), col("scale_lead_time3"))]

    for row in rows[hi + 1:]:
        if not row:
            continue
        first = _s(row[0]).lower()
        if first.startswith("sample"):   # skip the template's "Sample Row"
            continue
        part = _s(row[c_part]) if c_part is not None and c_part < len(row) else ""
        if not part:
            continue
        rr = _empty_reply_row()
        rr["Part Number"] = part
        any_scale = False
        for i, (cq, cp, cl) in enumerate(tiers, start=1):
            q = _num(row[cq]) if cq is not None and cq < len(row) else None
            p = _num(row[cp]) if cp is not None and cp < len(row) else None
            if q not in (None, 0):
                rr[f"Qty {i}"] = q
                any_scale = True
            if p not in (None, 0):
                rr[f"Price {i}"] = p
            if i == 1 and cl is not None and cl < len(row):
                lt = _num(row[cl])
                if lt not in (None, 0):
                    rr["L/T"] = lt
        # Boeing SAP scale qtys are blank until quoted; fall back to RFQ_QUANTITY as Qty 1
        if not any_scale and c_rfqqty is not None and c_rfqqty < len(row):
            rr["Qty 1"] = _num(row[c_rfqqty])
        res.rows.append(rr)
    res.customer_guess = "Boeing Distribution Services"
    res.warnings.append(
        "Boeing SAP template: quantity breaks are usually specified in the PDF's "
        "'Item Text' (QUOTE X AND Y); the .xlsx only carries RFQ_QUANTITY."
    )


def _adapt_boeing_pdf(content: bytes, res: NormalizeResult) -> None:
    import pdfplumber

    line_re = re.compile(r"^(\d{5})\s+([\d,]+)\s+(\S.*)$")
    quote_re = re.compile(r"QUOTE\s+([\d,]+)\s+AND\s+([\d,]+)", re.I)

    parts: list[dict] = []          # {line, qty_req, part}
    quote_qtys: list[tuple] = []     # aligned to item order via "QUOTE X AND Y"
    with pdfplumber.open(io.BytesIO(content)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for ln in text.split("\n"):
                ln = ln.strip()
                m = line_re.match(ln)
                if m:
                    part_field = m.group(3).split()[0]
                    parts.append({"line": m.group(1),
                                  "qty_req": _num(m.group(2)),
                                  "part": part_field})
                q = quote_re.search(ln)
                if q:
                    quote_qtys.append((_num(q.group(1)), _num(q.group(2))))

    for i, p in enumerate(parts):
        rr = _empty_reply_row()
        rr["Part Number"] = p["part"]
        if i < len(quote_qtys):
            rr["Qty 1"], rr["Qty 2"] = quote_qtys[i]
        else:
            rr["Qty 1"] = p["qty_req"]
        res.rows.append(rr)
    res.customer_guess = "Boeing Distribution Services"
    if not parts:
        res.warnings.append("Could not read line items from the PDF text layer (scanned PDF?).")
    res.warnings.append(
        "Boeing PDF: quantities taken from each item's 'QUOTE X AND Y' text; "
        "MFG/REV columns are present on the RFQ but blank until quoted."
    )


# --------------------------------------------------------------------------- #
# Generic parser + prediction (for unknown formats)
# --------------------------------------------------------------------------- #
# keyword hints -> reply column
_KEYWORDS = {
    "Part Number": ["part", "material", "item number", "adept", "mpn", "nsn", "cage part"],
    "Qty 1": ["qty 1", "quantity 1", "qty1", "quote qty 1", "scale_qty1", "break 1", "qty"],
    "Price 1": ["price 1", "cost 1", "unit price", "price1", "scale_price1", "price"],
    "Qty 2": ["qty 2", "quantity 2", "qty2", "quote qty 2", "scale_qty2", "break 2"],
    "Price 2": ["price 2", "cost 2", "price2", "scale_price2"],
    "Qty 3": ["qty 3", "quantity 3", "qty3", "quote qty 3", "scale_qty3", "break 3"],
    "Price 3": ["price 3", "cost 3", "price3", "scale_price3"],
    "L/T": ["l/t", "lead time", "leadtime", "delivery", "del", "lead"],
    "MFG": ["mfg", "manufacturer", "mfr", "cage", "make"],
    "REV": ["rev", "revision"],
}


def _detect_header_row(rows: list[list]) -> int:
    best_i, best_score = 0, -1.0
    for i, row in enumerate(rows[:12]):
        vals = [c for c in row if _s(c)]
        if not vals:
            continue
        strings = sum(1 for c in vals if isinstance(c, str) and not _s(c).replace(".", "").isdigit())
        score = len(vals) + strings * 0.5
        if score > best_score:
            best_score, best_i = score, i
    return best_i


def _suggest(source_columns: list[str]) -> dict:
    """Predict reply_col -> ranked source columns using keywords + fuzzy match."""
    out: dict[str, list[dict]] = {}
    for reply_col in REPLY_COLUMNS:
        kws = _KEYWORDS.get(reply_col, [])
        matches = []
        for src in source_columns:
            sl = src.lower().strip()
            score = 0.0
            for kw in kws:
                if sl == kw:
                    score = max(score, 1.0)
                elif kw in sl or sl in kw:
                    score = max(score, 0.82)
            fz = max(
                fuzz.ratio(reply_col.lower(), sl),
                fuzz.partial_ratio(reply_col.lower(), sl),
                max((fuzz.partial_ratio(kw, sl) for kw in kws), default=0),
            ) / 100.0
            score = max(score, fz * 0.75)
            if score >= 0.45:
                matches.append({"source_col": src, "score": round(score, 2)})
        out[reply_col] = sorted(matches, key=lambda m: -m["score"])[:5]
    return out


def _generic_parse(filename: str, content: bytes, res: NormalizeResult) -> None:
    _, rows = _load_sheet_matrix(content)
    if not rows:
        res.warnings.append("No data found in file.")
        return
    hi = _detect_header_row(rows)
    header = [_s(c) for c in rows[hi]]
    # de-duplicate / name blank headers
    seen = {}
    cols = []
    for j, h in enumerate(header):
        name = h or f"Column {j + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name} ({seen[name]})"
        else:
            seen[name] = 0
        cols.append(name)

    data_rows = rows[hi + 1:]
    records = []
    for row in data_rows:
        if not any(_s(c) for c in row):
            continue
        rec = {}
        for j, name in enumerate(cols):
            rec[name] = row[j] if j < len(row) else None
        records.append(rec)

    sample = {name: [] for name in cols}
    for rec in records[:5]:
        for name in cols:
            v = rec.get(name)
            if _s(v):
                sample[name].append(_s(v))

    res.source_columns = [c for c in cols if _s(c)]
    res.sample_data = sample
    res.candidate_rows = [[_s(c) for c in r] for r in rows[:10]]
    res.header_row_index = hi
    res.raw_records = records
    res.suggestions = _suggest(res.source_columns)


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #
def normalize(filename: str, content: bytes) -> NormalizeResult:
    fmt = detect_format(filename, content)
    res = NormalizeResult(filename=filename, detected_format=fmt, recognized=fmt != "unknown")
    try:
        if fmt == "boeing_pdf":
            _adapt_boeing_pdf(content, res)
        elif fmt in ("adept", "incora", "boeing_sap"):
            _, rows = _load_sheet_matrix(content)
            {"adept": _adapt_adept, "incora": _adapt_incora,
             "boeing_sap": _adapt_boeing_sap}[fmt](rows, res)
        else:
            _generic_parse(filename, content, res)
        if fmt != "unknown" and not res.rows:
            res.warnings.append("Recognized the format but found no part rows.")
    except Exception as exc:  # noqa: BLE001
        res.recognized = False
        res.detected_format = "unknown"
        res.warnings.append(f"Adapter failed ({exc}); falling back to manual mapping.")
        try:
            _generic_parse(filename, content, res)
        except Exception as exc2:  # noqa: BLE001
            res.warnings.append(f"Could not parse file: {exc2}")
    return res


def apply_mapping(raw_records: list[dict], mapping: dict) -> list[dict]:
    """mapping: reply_col -> source_col (or None). Build reply rows."""
    rows = []
    part_col = mapping.get("Part Number")
    for rec in raw_records:
        if part_col and not _s(rec.get(part_col)):
            continue
        rr = _empty_reply_row()
        for reply_col, src in mapping.items():
            if src and src in rec:
                rr[reply_col] = _num(rec[src]) if reply_col != "Part Number" else _s(rec[src])
        if any(_s(v) for v in rr.values()):
            rows.append(rr)
    return rows


def build_reply_workbook(rows: list[dict], sheet_name: str = "Quote") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    hdr_fill = PatternFill("solid", fgColor="234948")
    hdr_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.append(REPLY_COLUMNS)
    for j in range(1, len(REPLY_COLUMNS) + 1):
        c = ws.cell(1, j)
        c.fill, c.font, c.border = hdr_fill, hdr_font, border
        c.alignment = Alignment(horizontal="center")
    for rr in rows:
        ws.append([rr.get(c) for c in REPLY_COLUMNS])
    for j, name in enumerate(REPLY_COLUMNS, 1):
        width = 22 if name == "Part Number" else 10
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = width
    ws.freeze_panes = "A2"
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
